import openpyxl.workbook
import constant as const
import pandas as pd
import numpy  as np
import openpyxl
from openpyxl.styles import PatternFill



def create_random_data(rows, cols, output_excel):
    data=np.random.randint(-100,100,size=(rows,cols))
    df=pd.DataFrame(data)
    df.to_excel(output_excel,index=False)
    return df

def calculate_mean(df,output_excel):
    df['Row_mean'] = df.mean(axis=1)
    df.loc['Col_mean'] = df.mean(axis=0)
    df.to_excel(output_excel, index=False)
    return output_excel


def apply_color_base_value(cell, negative=const.RED_COLOR, positive=const.GREEN_COLOR):
    if cell.value <0:
        cell.fill=PatternFill(start_color=negative, end_color=negative,fill_type="solid")
    else :
        cell.fill=PatternFill(start_color=positive, end_color=positive,fill_type="solid")


def color_mean(file_path,output_excel,rows,cols):
    wb=openpyxl.load_workbook(file_path)
    ws=wb.active

    start_row = 2
    end_row = rows+2

    start_col = 1
    end_col = cols+1

    for row in range(start_row, end_row):
        cell=ws.cell(row=row, column=end_col)
        apply_color_base_value(cell)

    for col in range(start_col, end_col):
        cell=ws.cell(row=end_row, column=col)
        apply_color_base_value(cell)

    wb.save(output_excel)

def main():
    rows= const.ROWS
    cols= const.COLS
    df=create_random_data(rows, cols,'random_data.xlsx')
    file_path=calculate_mean(df,'mean_data.xlsx')
    color_mean(file_path,'mean_data_colored.xlsx',rows,cols)

if __name__ == "__main__":
    main()